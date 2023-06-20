import streamlit as st
import urllib.request
import time
#from dotenv import load_dotenv
from datetime import datetime,timedelta
import json
import urllib.request
import requests
import openai
import openpyxl

openai.api_key = st.secrets['OPENAI_API_KEY']

path = ""
user = path + "converse.xlsx"
wb_obj = openpyxl.load_workbook(user)
sheet_obj = wb_obj.active
looper=0
updatetext1=0
message = ""
healthmessage="This is health message of AI Chat BOT. It is running fine and you are receiving this message every 60 minutes "
healthtime = int(time.time())
api="9VCT9UPA8KX4M99P2U4I"
param1= "apikey=9VCT9UPA8KX4M99P2U4I&type=IN&getnotpulledonly=1"
param2="apikey=9VCT9UPA8KX4M99P2U4I&type=IN&markaspulled=1&getnotpulledonly=1"
polltime=2
while looper==0:
    cn=0

    updatefile = open(path + "updateid.txt", 'r+')
    updatetext = updatefile.read()
    updatefile.close()
    time.sleep(polltime)
    lastupdate = int(updatetext)
    try:
      f = urllib.request.urlopen('http://panel.rapiwha.com/get_messages.php/?'+ param1)
      if str(f.getcode())=="200":
          pass
    except requests.exceptions.RequestException as e:
        continue

    data=json.load(f)
    b=data
    c=str(b)
    print(c)
    lenc=len(c)
    if lenc==2:
        print("No New message "+str(datetime.now().strftime("%H:%M:%S")))

    if lenc > 2:
      f = urllib.request.urlopen('http://panel.rapiwha.com/get_messages.php/?' + param2)
      print("Message received at: " + str(datetime.now().strftime("%H:%M:%S")))
      tempstr = c[0:lenc]
      while cn==0:
       idpos=tempstr.find("id")
       numberpos = tempstr.find("from")
       messagepos = tempstr.find("text")
       datepos = tempstr.find("creation_date")
       if numberpos==-1:
        break
       numbertext = tempstr[numberpos + 10:messagepos - 40]
       print(numbertext)
       messageid=tempstr[idpos+6:idpos+15]
       messagetext=tempstr[messagepos+8:datepos-4]
       datetext=tempstr[datepos+17:datepos+36]
       datetimestr=datetext
       datetext=datetime.strptime(datetimestr,'%Y-%m-%d %H:%M:%S')
       datetext=datetext+timedelta(hours=8.5)
       datetext=str(datetext)
       print(messageid)
       print(datetext)
       print(messagetext)
       updatetext1 = messageid
       userfile = open(path + "user.txt", 'r', encoding='utf-8', errors='ignore')
       usertext=userfile.read()
       userfile.close()
       find1=usertext.find(numbertext)
       about = open(path + "companyinfo.txt", 'r', encoding='utf-8', errors='ignore')
       system_msg = about.read()
       about.close()
       print(len(system_msg))
       if find1==-1:
        userfile = open(path + "user.txt", 'a' ,encoding='utf-8', errors='ignore')
        userfile.write(numbertext+","+datetext+"\n")
        userfile.close()

       response1 = openai.ChatCompletion.create(model="gpt-3.5-turbo-0613",
                                               messages=[{"role": "system", "content": system_msg},
                                                         {"role": "user", "content": messagetext}])
       message1 = (response1["choices"][0]["message"]["content"])
       print(response1)
       data1= json.dumps(response1)
       b1 = data1

       c1 = str(b1)
       find1=c1.find("total_tokens")
       find2=c1.find("}}")
       total_t=c1[find1+15:find2]
       print(total_t)
       lenc1 = len(c1)
       print(c1)
       print(lenc1)
       print(message1)
       ab1 = 0
       roww1 = 2
       while ab1 == 0:

           cell_obj1 = sheet_obj.cell(row=roww1, column=1)
           cell_obj2 = sheet_obj.cell(row=roww1, column=2)
           cell_obj3 = sheet_obj.cell(row=roww1, column=3)
           cell_obj4 = sheet_obj.cell(row=roww1, column=4)
           cell_obj5 = sheet_obj.cell(row=roww1, column=5)

           if cell_obj1.value==None or cell_obj1.value=="":
               cell_obj1.value = numbertext
               cell_obj2.value = datetext
               cell_obj3.value = messagetext
               cell_obj4.value = int(total_t)
               cell_obj5.value = message1
               wb_obj.save(user)
               break
           else:
             roww1 = roww1 + 1
       print(len(message1))
       phonejn = "+91" + numbertext
       url = "https://panel.rapiwha.com/send_message.php"
       querystring = {"apikey": api, "number": phonejn,
                     "text": message1}
       response = requests.request("POST", url, params=querystring)
       print(response.text)
       find414 = response.text.find("414")
       if find414==-1:
          message1 = ""
          break

       else:
           text=message1
           splitter = 1000
           lent = len(text)
           s = 0
           a = 0
           g = 0
           while s <= lent:
               if text[s:s + 1] == " " and s > g + splitter:
                   text = text[:s] + '"#"' + text[s:]
                   a = s
                   g = g + splitter
               s = s + 1
           list = (text.split("#"))
           gtimes = 0
           while gtimes < len(list):
               gquery = list[gtimes]
               if gquery[0:1] != '"':
                   gquery = gquery.lstrip()
               if gquery[1:2] == ' ':
                   gquery = gquery[2:len(gquery)].lstrip()
               if gquery[len(gquery) - 1:len(gquery)] != '"':
                   gquery = gquery
               phonejn = "+91" + numbertext
               url = "https://panel.rapiwha.com/send_message.php"
               querystring = {"apikey": api, "number": phonejn,
                              "text": gquery}
               response = requests.request("POST", url, params=querystring)

               print(response.text)
               #message1 = ""
               #print(gquery)

               gtimes = gtimes + 1
           break

      tempstr = tempstr[datepos + 65:lenc]
    if int(updatetext1)>=int(updatetext):
      updatefile = open(path + "updateid.txt", 'w')
      updatetext = int(updatetext1)+1
      updatefile.write(str(updatetext))
      updatefile.close()
    healthtime1 = int(time.time())
    print(healthtime1)
    if healthtime1 - healthtime > 3600:
        phonejn = "+91" + "6398370244"
        url = "https://panel.rapiwha.com/send_message.php"
        querystring = {"apikey": api, "number": phonejn,
                       "text": healthmessage}
        response = requests.request("GET", url, params=querystring)
        #time.sleep(0.5)
        print(response.text)
        healthtime = healthtime1
    print("Offset ID in Text File updated succesfully")

